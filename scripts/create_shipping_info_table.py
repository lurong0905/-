#!/usr/bin/env python
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - handled at runtime for minimal installs
    PdfReader = None


INFO_TABLE_NAME = "发货信息搜集表.xlsx"

NUCLEAR_PLANTS = {
    "3010": {
        "name": "秦山核电有限公司（秦一厂）",
        "address": "浙江省海盐县秦山核电有限公司（一厂）现场物资仓库",
        "contact": "李兴珍",
        "phone": "13626730627",
        "email": "LIXZ@CNNP.COM.CN",
    },
    "3011": {
        "name": "秦山核电有限公司（方家山）",
        "address": "浙江省海盐县秦山核电有限公司（方家山）现场AB库",
        "contact": "徐莉莉",
        "phone": "13586356712",
        "email": "xulili@cnnp.com.cn",
    },
    "3020": {
        "name": "核电秦山联营有限公司（秦二厂）",
        "address": "浙江省海盐县核电秦山联营有限公司（二厂）现场物资仓库",
        "contact": "陈燕",
        "phone": "13626769903",
        "email": "chenyan01@cnnp.com.cn",
    },
    "3030": {
        "name": "秦山第三核电有限公司（秦三厂）",
        "address": "浙江省海盐县秦山第三核电有限公司（三厂）现场物资仓库",
        "contact": "步建军",
        "phone": "13515734297",
        "email": "bujj@cnnp.com.cn",
    },
    "3050": {
        "name": "江苏核电有限公司",
        "address": "江苏省连云港市连云区宿城乡核电南路9000号",
        "contact": "许剑",
        "phone": "18961371827",
        "email": "xujian02@cnnp.com.cn",
    },
    "3060": {
        "name": "三门核电有限公司",
        "address": "浙江省三门县三门核电有限公司永久仓库",
        "contact": "祁丽莎",
        "phone": "18857657584",
        "email": "qils@cnnp.com.cn",
    },
    "3070": {
        "name": "福建福清核电有限公司",
        "address": "福建省福清市三山镇福清核电有限公司",
        "contact": "魏富建",
        "phone": "17758936501 / 17758936510",
        "email": "",
    },
    "3080": {
        "name": "海南核电有限公司",
        "address": "海南省昌江县海尾镇核电厂EF仓库",
        "contact": "刘凯",
        "phone": "18976717520",
        "email": "t_liukai@cnnp.com.cn",
    },
    "3130": {
        "name": "中核国电漳州能源有限公司",
        "address": "福建省漳州市云霄县列屿镇城内村城东999号漳州核电现场AB库",
        "contact": "方泽琳；陈志川",
        "phone": "18695755693；15260064065",
        "email": "t_fangz102@cnnp.com.cn；t_chenzc03@cnnp.com.cn",
    },
    "3131": {
        "name": "中核国电漳州能源有限公司",
        "address": "福建省漳州市云霄县列屿镇城内村城东999号漳州核电现场AB库",
        "contact": "方泽琳；陈志川",
        "phone": "18695755693；15260064065",
        "email": "t_fangz102@cnnp.com.cn；t_chenzc03@cnnp.com.cn",
    },
    "3110": {
        "name": "中核辽宁核电有限公司",
        "address": "辽宁省兴城市滨海大道100号中核辽宁核电有限公司",
        "contact": "孙佳宾",
        "phone": "13258701728",
        "email": "sunjb@cnnp.com.cn",
    },
    "3111": {
        "name": "中核辽宁核电有限公司",
        "address": "辽宁省兴城市滨海大道100号中核辽宁核电有限公司",
        "contact": "孙佳宾",
        "phone": "13258701728",
        "email": "sunjb@cnnp.com.cn",
    },
    "3480": {
        "name": "中核山东核能有限公司",
        "address": "山东省烟台市海阳市中核山东核能有限公司",
        "contact": "李赤峰",
        "phone": "13027520431",
        "email": "licf@cnnp.com.cn",
    },
    "3270": {
        "name": "中核龙原科技有限公司",
        "address": "福建省宁德市霞浦县长春镇长门村长表岛核电现场仓储科接货组",
        "contact": "钟毅",
        "phone": "17689660090",
        "email": "zhongyi@cnnp.com.cn",
    },
}

BUYER_TO_PLANT_CODE = {item["name"]: code for code, item in NUCLEAR_PLANTS.items()}

FIXED_FIELDS = [
    "客户体系",
    "合同名称",
    "合同号",
    "买方名称",
    "卖方名称",
    "工厂代码",
    "收货单位",
    "收货地址",
    "收货人",
    "收货电话",
    "收货邮箱",
    "合同管理联系人",
    "项目管理联系人",
    "合同要求交货时间",
    "发货地点",
    "发货单位",
    "发货单位联系人",
    "发货单位联系电话",
    "立项单号默认值",
    "采购申请处室项目负责人默认值",
    "备注",
]

BATCH_HEADERS = [
    "批次",
    "发货清单文件",
    "计划发运时间",
    "预计到达时间",
    "运单号/快递单号",
    "运输方式",
    "装卸货方式",
    "储存级别",
    "储存条件",
    "是否生成最终版",
    "备注",
]

REQUIRED_BATCH_FIELDS = {"计划发运时间", "装卸货方式", "储存级别", "储存条件"}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[ \t]+", " ", str(value)).strip()


def first_match(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I | re.M)
        if match:
            return clean_text(match.group(1))
    return ""


def discover_contract_pdf(folder: Path) -> Path | None:
    pdfs = [p for p in folder.glob("*.pdf") if not p.name.startswith("~$")]
    if not pdfs:
        return None
    preferred = [p for p in pdfs if "合同" in p.stem]
    return sorted(preferred or pdfs, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def discover_shipment_file(folder: Path) -> Path | None:
    candidates = [
        p
        for p in folder.glob("*.xls*")
        if not p.name.startswith("~$") and p.name != INFO_TABLE_NAME
    ]
    if not candidates:
        return None
    preferred = [p for p in candidates if any(key in p.stem for key in ("发货清单", "装箱清单", "装箱"))]
    return sorted(preferred or candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def extract_pdf_text(pdf_path: Path | None) -> str:
    if not pdf_path or PdfReader is None:
        return ""
    try:
        reader = PdfReader(str(pdf_path))
    except Exception:
        return ""
    total_pages = len(reader.pages)
    page_indexes = list(range(min(30, total_pages)))
    page_indexes.extend(range(max(0, total_pages - 12), total_pages))
    page_indexes = sorted(set(page_indexes))
    parts: list[str] = []
    for index in page_indexes:
        try:
            parts.append(reader.pages[index].extract_text() or "")
        except Exception:
            continue
    return "\n".join(parts)


def infer_contract_name(folder: Path, pdf_path: Path | None, text: str) -> tuple[str, str]:
    if "合同" in folder.name:
        return folder.name, "合同文件夹名称"
    if pdf_path and "合同" in pdf_path.stem:
        return pdf_path.stem, "PDF 文件名"
    value = first_match(text, [r"合同名称[：:\s]*([^\n\r]+)", r"采购合同名称[：:\s]*([^\n\r]+)"])
    return value, "PDF 预填" if value else "需用户填写"


def infer_contract_no(text: str) -> str:
    return first_match(
        text,
        [
            r"\b([A-Z]{2,}\d*[A-Z0-9]*-\d{6,}-\d{2,})\b",
            r"(?:采购合同号|合同编号|合同号)[：:\s]*([A-Z0-9][A-Z0-9/_-]{5,})",
        ],
    )


def infer_party(text: str, label: str) -> str:
    if label.startswith("买"):
        for company in sorted(BUYER_TO_PLANT_CODE, key=len, reverse=True):
            if company in text:
                return company
    if label.startswith("卖") and "江苏道众能源科技有限公司" in text:
        return "江苏道众能源科技有限公司"
    spaced_label = r"\s*".join(label)
    value = first_match(text, [rf"{spaced_label}[：:\s]*([^\n\r，,；;]+)"])
    if value and ("公司" in value or "厂" in value) and ")" not in value and "）" not in value:
        return value
    return ""


def infer_contact(text: str, label: str) -> str:
    return first_match(
        text,
        [
            rf"{label}[：:\s]*([\u4e00-\u9fffA-Za-z·]+)",
            rf"{label}\s*[:：]\s*([^\n\r，,；; ]+)",
        ],
    )


def infer_delivery_time(text: str) -> str:
    compact = re.sub(r"\s+", "", text)
    if "交货时间详见供货清单" in compact or re.search(r"交货时间.{0,12}详见.{0,4}供货清单", compact):
        return "详见供货清单"
    value = first_match(text, [r"合同要求交货时间[：:\s]*([^\n\r。；;]+)", r"交货时间[：:\s]*([^\n\r。；;]+)"])
    if "…" in value or value.strip().isdigit():
        return ""
    return value


def infer_factory_code(text: str, buyer: str) -> str:
    for name, code in BUYER_TO_PLANT_CODE.items():
        if name and (name == buyer or name in text):
            return code
    value = first_match(text, [r"工厂代码[：:\s]*(\d{4})", r"工厂[：:\s]*(\d{4})"])
    if value in NUCLEAR_PLANTS:
        return value
    for code in NUCLEAR_PLANTS:
        if re.search(rf"\b{code}\b", text):
            return code
    return ""


def infer_batch_name(shipment_file: Path | None) -> str:
    if not shipment_file:
        return "第一批"
    match = re.search(r"第[一二三四五六七八九十0-9]+批", shipment_file.stem)
    return match.group(0) if match else "第一批"


def build_facts(folder: Path) -> tuple[dict[str, str], dict[str, str], Path | None, Path | None]:
    pdf_path = discover_contract_pdf(folder)
    shipment_file = discover_shipment_file(folder)
    text = extract_pdf_text(pdf_path)

    facts: dict[str, str] = {}
    sources: dict[str, str] = {}

    contract_name, contract_name_source = infer_contract_name(folder, pdf_path, text)
    facts["合同名称"] = contract_name
    sources["合同名称"] = contract_name_source

    contract_no = infer_contract_no(text)
    facts["合同号"] = contract_no
    sources["合同号"] = "PDF 预填" if contract_no else "需用户填写"

    buyer = infer_party(text, "买方")
    seller = infer_party(text, "卖方")
    facts["买方名称"] = buyer
    facts["卖方名称"] = seller
    sources["买方名称"] = "PDF 预填" if buyer else "需用户填写"
    sources["卖方名称"] = "PDF 预填" if seller else "需用户填写"

    factory_code = infer_factory_code(text, buyer)
    facts["工厂代码"] = factory_code
    sources["工厂代码"] = "PDF/买方匹配预填" if factory_code else "需用户填写"

    customer_system = "中核" if factory_code in NUCLEAR_PLANTS or "中核" in text else ""
    facts["客户体系"] = customer_system
    sources["客户体系"] = "工厂代码匹配" if customer_system else "需用户填写"

    plant = NUCLEAR_PLANTS.get(factory_code, {})
    facts["收货单位"] = buyer or plant.get("name", "")
    facts["收货地址"] = plant.get("address", "")
    facts["收货人"] = plant.get("contact", "")
    facts["收货电话"] = plant.get("phone", "")
    facts["收货邮箱"] = plant.get("email", "")
    for field in ("收货单位", "收货地址", "收货人", "收货电话", "收货邮箱"):
        sources[field] = "Skill 内置中核默认收货地址" if facts.get(field) else "需用户填写"

    facts["合同管理联系人"] = infer_contact(text, "合同管理联系人")
    facts["项目管理联系人"] = infer_contact(text, "项目管理联系人")
    sources["合同管理联系人"] = "PDF 预填" if facts["合同管理联系人"] else "需用户填写"
    sources["项目管理联系人"] = "PDF 预填" if facts["项目管理联系人"] else "可留空，仅记录合同事实"

    facts["合同要求交货时间"] = infer_delivery_time(text)
    sources["合同要求交货时间"] = "PDF 预填" if facts["合同要求交货时间"] else "需用户填写"

    seller_defaults = seller == "江苏道众能源科技有限公司"
    facts["发货地点"] = "江苏省连云港市海州区昌意路联东U谷A5-2#" if seller_defaults else ""
    facts["发货单位"] = seller
    facts["发货单位联系人"] = ""
    facts["发货单位联系电话"] = "0518-81180066" if seller_defaults else ""
    sources["发货地点"] = "卖方默认地址" if facts["发货地点"] else "需用户填写"
    sources["发货单位"] = "卖方名称" if seller else "需用户填写"
    sources["发货单位联系人"] = "需用户填写"
    sources["发货单位联系电话"] = "卖方默认电话" if facts["发货单位联系电话"] else "需用户填写"

    facts["立项单号默认值"] = "/"
    facts["采购申请处室项目负责人默认值"] = "/"
    facts["备注"] = ""
    sources["立项单号默认值"] = "中核模板默认"
    sources["采购申请处室项目负责人默认值"] = "中核模板默认"
    sources["备注"] = ""

    return facts, sources, pdf_path, shipment_file


def set_common_styles(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    ws.freeze_panes = "A2"


def add_fixed_sheet(wb: Workbook, facts: dict[str, str], sources: dict[str, str]) -> None:
    ws = wb.active
    ws.title = "合同固定信息"
    ws.append(["字段", "值", "来源/说明"])
    required_blank_fill = PatternFill("solid", fgColor="FFF2CC")
    for field in FIXED_FIELDS:
        ws.append([field, facts.get(field, ""), sources.get(field, "")])
        if field in {"合同号", "买方名称", "卖方名称", "收货地址", "收货人"} and not facts.get(field):
            ws.cell(ws.max_row, 2).fill = required_blank_fill
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 34
    dv = DataValidation(type="list", formula1='"中核,国核,广核,其他"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B2")
    set_common_styles(ws)


def add_batch_sheet(wb: Workbook, shipment_file: Path | None) -> None:
    ws = wb.create_sheet("批次发货信息")
    ws.append(BATCH_HEADERS)
    batch = infer_batch_name(shipment_file)
    ws.append([
        batch,
        shipment_file.name if shipment_file else "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "否",
        "必确认字段未填全时只能生成待确认版",
    ])
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    for col_index, header in enumerate(BATCH_HEADERS, start=1):
        if header in REQUIRED_BATCH_FIELDS:
            ws.cell(2, col_index).fill = required_fill
            ws.cell(2, col_index).comment = Comment("生成最终文件前必须填写或由用户确认。", "Codex")
    widths = [14, 32, 18, 18, 20, 16, 28, 14, 16, 16, 36]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for cell_range, values in {
        "J2:J200": '"是,否"',
        "I2:I200": '"室内,室外,其它"',
    }.items():
        dv = DataValidation(type="list", formula1=values, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(cell_range)
    set_common_styles(ws)


def add_help_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("填写说明与选项")
    ws.append(["字段", "填写说明", "允许值/示例"])
    rows = [
        ("计划发运时间", "必须由用户填写或确认，不得从合同、模板、发货清单或当前日期推断。", "2026年6月1日"),
        ("预计到达时间", "资料不明确时必须由用户填写或确认。", "2026年6月3日"),
        ("装卸货方式", "可多选，多个值用中文分号分隔。", "吊车；叉车；人工搬运；其它"),
        ("储存级别", "按合同或物项要求填写具体等级。", "A/B/C/D 或合同指定等级"),
        ("储存条件", "在室内、室外、其它中选择；其它需补充说明。", "室内"),
        ("是否生成最终版", "必确认字段不全时填否，只生成待确认版。", "是/否"),
        ("商务合同处合同经理", "中核模板取合同固定信息中的合同管理联系人。", "刘原序"),
        ("立项单号", "中核模板一般不填，默认使用合同固定信息中的 `/`。", "/"),
        ("采购申请处室项目负责人", "中核模板一般不填，默认使用合同固定信息中的 `/`。", "/"),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 34
    set_common_styles(ws)


def create_workbook(output_path: Path, facts: dict[str, str], sources: dict[str, str], shipment_file: Path | None) -> None:
    wb = Workbook()
    add_fixed_sheet(wb, facts, sources)
    add_batch_sheet(wb, shipment_file)
    add_help_sheet(wb)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="创建或预填发货信息搜集表。")
    parser.add_argument("contract_folder", nargs="?", default=".", help="合同文件夹路径，默认当前目录。")
    parser.add_argument("--output", help=f"输出文件路径，默认合同文件夹内的 {INFO_TABLE_NAME}。")
    parser.add_argument("--overwrite", action="store_true", help="覆盖已存在的信息搜集表。")
    args = parser.parse_args()

    folder = Path(args.contract_folder).resolve()
    if not folder.exists() or not folder.is_dir():
        raise SystemExit(f"合同文件夹不存在：{folder}")

    output_path = Path(args.output).resolve() if args.output else folder / INFO_TABLE_NAME
    if output_path.exists() and not args.overwrite:
        print(json.dumps({"status": "exists", "path": str(output_path)}, ensure_ascii=False, indent=2))
        return 0

    facts, sources, pdf_path, shipment_file = build_facts(folder)
    create_workbook(output_path, facts, sources, shipment_file)

    missing_fixed = [
        field
        for field in ("合同号", "买方名称", "卖方名称", "收货地址", "收货人", "发货单位联系人")
        if not facts.get(field)
    ]
    result = {
        "status": "created",
        "path": str(output_path),
        "pdf": str(pdf_path) if pdf_path else "",
        "shipment_file": shipment_file.name if shipment_file else "",
        "missing_fixed_fields": missing_fixed,
        "required_batch_fields": sorted(REQUIRED_BATCH_FIELDS),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
